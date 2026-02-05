import gradio as gr
import pandas as pd
import numpy as np
import cv2
import os
import json

# --- 全域變數:預設標準 ---
DEFAULT_STANDARDS = {
    # 檢查開關
    "check_parameter_1": True,
    "check_parameter_2": True,
    "check_parameter_3": True,
    "check_image": True,
    
    # 參數值
    "parameter_1_min": 20.0,
    "parameter_1_pass": 50.0,
    "parameter_2_min": 1.8,
    "parameter_2_max": 2.0,
    "parameter_2_pass": 1.9,
    "parameter_3_min": 2.0,
    "parameter_3_pass": 2.2,
    "image_threshold_1": 50,
    "image_threshold_2": 100,
    
    # 單位設定
    "parameter_1_unit": "ng/μL",
    "parameter_2_unit": "ratio",
    "parameter_3_unit": "ratio",
    "image_unit": "intensity"
}

# 當前使用的標準(會被使用者修改)
CURRENT_STANDARDS = DEFAULT_STANDARDS.copy()


# --- 標準管理函式 ---
def update_standards(check_p1, check_p2, check_p3, check_img,
                     p1_min, p1_pass, p1_unit,
                     p2_min, p2_max, p2_pass, p2_unit,
                     p3_min, p3_pass, p3_unit,
                     img_t1, img_t2, img_unit):
    """更新判定標準"""
    global CURRENT_STANDARDS
    CURRENT_STANDARDS = {
        "check_parameter_1": bool(check_p1),
        "check_parameter_2": bool(check_p2),
        "check_parameter_3": bool(check_p3),
        "check_image": bool(check_img),
        "parameter_1_min": float(p1_min),
        "parameter_1_pass": float(p1_pass),
        "parameter_1_unit": str(p1_unit),
        "parameter_2_min": float(p2_min),
        "parameter_2_max": float(p2_max),
        "parameter_2_pass": float(p2_pass),
        "parameter_2_unit": str(p2_unit),
        "parameter_3_min": float(p3_min),
        "parameter_3_pass": float(p3_pass),
        "parameter_3_unit": str(p3_unit),
        "image_threshold_1": float(img_t1),
        "image_threshold_2": float(img_t2),
        "image_unit": str(img_unit)
    }
    return "Standards updated successfully!"

def reset_standards():
    """重置為預設標準"""
    global CURRENT_STANDARDS
    CURRENT_STANDARDS = DEFAULT_STANDARDS.copy()
    return (
        DEFAULT_STANDARDS["check_parameter_1"],
        DEFAULT_STANDARDS["check_parameter_2"],
        DEFAULT_STANDARDS["check_parameter_3"],
        DEFAULT_STANDARDS["check_image"],
        DEFAULT_STANDARDS["parameter_1_min"],
        DEFAULT_STANDARDS["parameter_1_pass"],
        DEFAULT_STANDARDS["parameter_1_unit"],
        DEFAULT_STANDARDS["parameter_2_min"],
        DEFAULT_STANDARDS["parameter_2_max"],
        DEFAULT_STANDARDS["parameter_2_pass"],
        DEFAULT_STANDARDS["parameter_2_unit"],
        DEFAULT_STANDARDS["parameter_3_min"],
        DEFAULT_STANDARDS["parameter_3_pass"],
        DEFAULT_STANDARDS["parameter_3_unit"],
        DEFAULT_STANDARDS["image_threshold_1"],
        DEFAULT_STANDARDS["image_threshold_2"],
        DEFAULT_STANDARDS["image_unit"],
        "Standards reset to default!"
    )

def save_standards_preset(preset_name):
    """儲存當前標準為預設檔"""
    if not preset_name:
        return None, "Please enter a preset name"
    
    preset_file = f"{preset_name}_standards.json"
    with open(preset_file, 'w') as f:
        json.dump(CURRENT_STANDARDS, f, indent=2)
    
    return preset_file, f"Preset '{preset_name}' saved successfully!"

def load_standards_preset(preset_file):
    """載入預設檔"""
    global CURRENT_STANDARDS
    if preset_file is None:
        return (None,)*17 + ("Please select a preset file",)
    
    try:
        with open(preset_file.name, 'r') as f:
            CURRENT_STANDARDS = json.load(f)
        
        return (
            CURRENT_STANDARDS["check_parameter_1"],
            CURRENT_STANDARDS["check_parameter_2"],
            CURRENT_STANDARDS["check_parameter_3"],
            CURRENT_STANDARDS["check_image"],
            CURRENT_STANDARDS["parameter_1_min"],
            CURRENT_STANDARDS["parameter_1_pass"],
            CURRENT_STANDARDS["parameter_1_unit"],
            CURRENT_STANDARDS["parameter_2_min"],
            CURRENT_STANDARDS["parameter_2_max"],
            CURRENT_STANDARDS["parameter_2_pass"],
            CURRENT_STANDARDS["parameter_2_unit"],
            CURRENT_STANDARDS["parameter_3_min"],
            CURRENT_STANDARDS["parameter_3_pass"],
            CURRENT_STANDARDS["parameter_3_unit"],
            CURRENT_STANDARDS["image_threshold_1"],
            CURRENT_STANDARDS["image_threshold_2"],
            CURRENT_STANDARDS["image_unit"],
            "Preset loaded successfully!"
        )
    except Exception as e:
        return (None,)*17 + (f"Error loading preset: {str(e)}",)


def analyze_custom_standards_data(file_objs):
    """在自訂標準頁面分析多個 Excel 檔案"""
    if not file_objs:
        return None, None, "Please upload Excel files"
    
    all_results = []
    
    # 取得當前標準和單位
    p1_min = CURRENT_STANDARDS["parameter_1_min"]
    p1_pass = CURRENT_STANDARDS["parameter_1_pass"]
    p1_unit = CURRENT_STANDARDS["parameter_1_unit"]
    
    p2_min = CURRENT_STANDARDS["parameter_2_min"]
    p2_max = CURRENT_STANDARDS["parameter_2_max"]
    p2_pass = CURRENT_STANDARDS["parameter_2_pass"]
    p2_unit = CURRENT_STANDARDS["parameter_2_unit"]
    
    p3_min = CURRENT_STANDARDS["parameter_3_min"]
    p3_pass = CURRENT_STANDARDS["parameter_3_pass"]
    p3_unit = CURRENT_STANDARDS["parameter_3_unit"]
    
    check_p1 = CURRENT_STANDARDS["check_parameter_1"]
    check_p2 = CURRENT_STANDARDS["check_parameter_2"]
    check_p3 = CURRENT_STANDARDS["check_parameter_3"]
    
    for f in file_objs:
        try:
            # 用 header=0 讀取 demo 檔案格式
            df_raw = pd.read_excel(f.name, header=0)
            
            for i in range(len(df_raw)):
                try:
                    sample_name = str(df_raw.iloc[i, 1])
                    param1 = round(float(df_raw.iloc[i, 2]))
                    param2 = float(df_raw.iloc[i, 3])
                    param3 = float(df_raw.iloc[i, 4])
                    
                    issues = []
                    
                    # 根據開關檢查各參數
                    if check_p1 and param1 < p1_min:
                        issues.append(f"Parameter 1 < {p1_min} {p1_unit}")
                    
                    if check_p2:
                        if param2 < p2_min:
                            issues.append(f"Parameter 2 < {p2_min} {p2_unit}")
                        if param2 > p2_max:
                            issues.append(f"Parameter 2 > {p2_max} {p2_unit}")
                    
                    if check_p3 and param3 < p3_min:
                        issues.append(f"Parameter 3 < {p3_min} {p3_unit}")
                    
                    # 判定結果
                    if len(issues) > 0:
                        quality = 'FAIL'
                        note = '; '.join(issues)
                    else:
                        # 檢查是否符合標準
                        pass_conditions = []
                        if check_p1:
                            pass_conditions.append(param1 >= p1_pass)
                        if check_p2:
                            pass_conditions.append(param2 >= p2_pass)
                        if check_p3:
                            pass_conditions.append(param3 >= p3_pass)
                        
                        if all(pass_conditions):
                            quality = 'PASS'
                            note = 'Excellent quality'
                        else:
                            quality = 'ACCEPTABLE'
                            note = 'Meets minimum requirements'
                    
                    result_row = {
                        'Sample Name': sample_name,
                        f'Parameter 1 ({p1_unit})': param1,
                        f'Parameter 2 ({p2_unit})': param2,
                        f'Parameter 3 ({p3_unit})': param3,
                        'QC': quality,
                        'Note': note
                    }
                    all_results.append(result_row)
                    
                except Exception as e:
                    try:
                        sample_name = str(df_raw.iloc[i, 1])
                    except:
                        sample_name = 'Unknown'
                    
                    all_results.append({
                        'Sample Name': sample_name,
                        f'Parameter 1 ({p1_unit})': 'ERROR',
                        f'Parameter 2 ({p2_unit})': 'ERROR',
                        f'Parameter 3 ({p3_unit})': 'ERROR',
                        'QC': 'ERROR',
                        'Note': f'Cannot read values: {str(e)}'
                    })
        except Exception as e:
            all_results.append({
                'Sample Name': f'File Error: {os.path.basename(f.name)}',
                f'Parameter 1 ({p1_unit})': 'ERROR',
                f'Parameter 2 ({p2_unit})': 'ERROR',
                f'Parameter 3 ({p3_unit})': 'ERROR',
                'QC': 'ERROR',
                'Note': f'Cannot read file: {str(e)}'
            })
            continue
    
    if not all_results:
        return None, None, "Failed to process any files"
    
    # 建立 DataFrame
    result_df = pd.DataFrame(all_results)
    
    # 儲存為 Excel (使用你的 generate_qc_report 函數)
    output_path = os.path.abspath("(Custom)QC_Analysis_Report.xlsx")
    generate_qc_report(result_df, output_path)
    
    success_msg = f"Analysis completed! Processed {len(all_results)} samples."
    
    return result_df, output_path, success_msg

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font

def generate_qc_report(df, file_name="(Custom)QC_Analysis_Report.xlsx"):
    # 建立活頁簿與工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis Results"

    # 定義樣式
    blue_fill = PatternFill(start_color="598BAB", fill_type="solid")  # 標題用藍色
    green_fill = PatternFill(start_color="CDFFCF", fill_type="solid")  # 合格
    yellow_fill = PatternFill(start_color="FFFF99", fill_type="solid")  # 可接受
    red_fill = PatternFill(start_color="FFA3A3", fill_type="solid")  # 不合格
    
    # 寫入 DataFrame 到工作表
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # 標題列上色 (第一列)
    for cell in ws[1]:  
        cell.fill = blue_fill
        cell.font = Font(bold=True, color="e7cd79")  # 白色粗體字
    
    # 找出需要處理的欄位索引
    try:
        ratio_280_col = df.columns.get_loc("260/280") + 1  # +1 因為 openpyxl 從 1 開始
    except KeyError:
        ratio_280_col = None
        print("找不到 260/280 欄位")
    
    try:
        ratio_230_col = df.columns.get_loc("260/230") + 1
    except KeyError:
        ratio_230_col = None
        print("找不到 260/230 欄位")
    
    try:
        status_col = df.columns.get_loc("QC") + 1
    except KeyError:
        status_col = None
        print("找不到 Status 欄位")
    
    # 從第二列開始處理資料 (跳過標題)
    for row_idx in range(2, ws.max_row + 1):
        
        # 處理 260/280 欄位上色
        if ratio_280_col:
            ratio_280_cell = ws.cell(row=row_idx, column=ratio_280_col)
            ratio_280_value = ratio_280_cell.value
            
            if ratio_280_value is not None:
                try:
                    ratio_280_value = float(ratio_280_value)
                    if 2.0 <= ratio_280_value <= 2.2:
                        ratio_280_cell.fill = green_fill
                    else:
                        ratio_280_cell.fill = red_fill
                except (ValueError, TypeError):
                    ratio_280_cell.fill = red_fill
        
        # 處理 260/230 欄位上色
        if ratio_230_col:
            ratio_230_cell = ws.cell(row=row_idx, column=ratio_230_col)
            ratio_230_value = ratio_230_cell.value
            
            if ratio_230_value is not None:
                try:
                    ratio_230_value = float(ratio_230_value)
                    if ratio_230_value >= 1.5:
                        ratio_230_cell.fill = green_fill
                    else:
                        ratio_230_cell.fill = red_fill
                except (ValueError, TypeError):
                    ratio_230_cell.fill = red_fill
            else:
                ratio_230_cell.fill = red_fill
        
        # 處理 Status 欄位上色
        if status_col:
            status_cell = ws.cell(row=row_idx, column=status_col)
            status_value = str(status_cell.value).upper() if status_cell.value else ""
            
            if status_value == "PASS":
                status_cell.fill = green_fill
            elif status_value == "ACCEPTABLE":
                status_cell.fill = yellow_fill
            elif status_value == "FAIL":
                status_cell.fill = red_fill

    # 設定欄寬
    for column in ws.columns:
        ws.column_dimensions[column[0].column_letter].width = 20

    # 儲存檔案
    wb.save(file_name)
    print(f" Report saved as {file_name}")

#笑到崩潰直接寫到亂倒一鍋粥,那這碗粥就直接claude喝下去吧

# --- 1. Gel Image Analysis Logic ---
def analyze_gel_image(image_path, lane_index, total_lanes=14):
    """
    電泳影像分析函式
    功能:分析電泳圖中特定 Lane 的品質
    參數:
        - image_path: 影像檔案路徑
        - lane_index: 要分析的 Lane 編號 (從 0 開始)
        - total_lanes: 總共有幾條 Lane (預設 14)
    回傳:(smear_status, integrity_score, n_result)
    """
    if image_path is None:
        return "No Image", "N/A", "4"
    
    # 備註:讀取影像為灰階格式
    img = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
    
    if img is None:
        return "Read Error", "N/A", "4"

    # 黑白反轉邏輯 - 平均亮度 > 127 代表背景是白色,需反轉
    if np.mean(img) > 127:
        img = 255 - img

    # 計算每條 Lane 的寬度並切割目標區域
    h, w = img.shape
    lane_w = w // total_lanes
    start_x = lane_index * lane_w
    lane_roi = img[:, start_x : start_x + lane_w]
    
    # 計算平均亮度用於判斷拖尾
    avg_brightness = np.mean(lane_roi)
    
    # 初步判斷是否有 Smearing(拖尾現象)
    # 門檻值 50 可依樣本特性調整
    if avg_brightness > 50:
        smear_status = "Smearing"
    else:
        smear_status = "Clean"

    # 偵測三個標記區域的亮度
    # 這些比例 (0.15, 0.25 等) 需依實際 Ladder 位置調整
    bright_20k = np.max(lane_roi[int(h*0.15):int(h*0.25), :])  # 20kb 區域
    bright_5k  = np.max(lane_roi[int(h*0.45):int(h*0.55), :])  # 5kb 區域
    bright_3k  = np.max(lane_roi[int(h*0.65):int(h*0.75), :])  # 3kb 區域
    
    # 複雜的 Smearing 狀態判定邏輯
    # 這部分根據各區域亮度關係來細化拖尾判斷
    if smear_status == "Smearing":
        if avg_brightness > bright_3k:
            smear_status = "smearing"
        else:
            smear_status = "smear"
            
        if avg_brightness < bright_5k:
            if "3k" in str(bright_5k):
                smear_status = "smearing"
            else:
                smear_status = ""
                
        if avg_brightness < bright_20k:
            if "5k" in str(bright_20k):
                smear_status = "smearing"
            else:
                smear_status = ""
                
        if avg_brightness > bright_20k:
            if "visible" in str(bright_20k):
                smear_status = "smearing"
            else:
                smear_status = ""

    # 條帶完整度判定
    if smear_status != "Smear":
        status = ""
        integrity_score = "Low"
        
        # 亮度 > 100 視為可見條帶
        # 門檻值 100 可依需求調整
        if bright_20k > 100:
            status = "band integrity"
            integrity_score = "Visible"
        elif bright_5k > 100 or bright_3k > 100:
            status = "band accptable"
            integrity_score = "Medium"
        else:
            status = "No Band"
            integrity_score = "N/A"
    else:
        status = "unqualified"
        integrity_score = "Low"

    # 備註:
    #綜合判定品質等級 (1-4)
    # 1 = 最優,4 = 最差
    if smear_status and integrity_score == "Visible":
        n_result = "1"
    elif (smear_status == "5k" and integrity_score == "Visible") or status == "band accptable":
        n_result = "2"
    elif (smear_status == "3k" and integrity_score == "Visible") or status == "band accptable":
        n_result = "3"
    else:
        n_result = "4"

    return smear_status, integrity_score, n_result


def style_dataframe(df):
    """
    表格顏色標註函式
    功能:根據品質判定結果上色,方便快速辨識
    """
    def color_rows(row):
        if 'Quality Check' not in row.index:
            return [''] * len(row)
        
        quality = row['Quality Check']
        
        if quality == 'PASS':
            return ['background-color: #90EE90'] * len(row)
        elif quality == 'ACCEPTABLE':
            return ['background-color: #87CEEB'] * len(row)
        elif quality == 'FAIL':
            return ['background-color: #FFB6C6'] * len(row)
        elif quality == 'ERROR':
            return ['background-color: #D3D3D3'] * len(row)
        else:
            return [''] * len(row)
    
    return df.style.apply(color_rows, axis=1)


# --- 2. Stunner Data Loading with Color Annotation ---
def load_single_stunner(file_obj):
    """
    載入單一 Stunner 檔案
    功能:處理單一檔案上傳並分析品質
    """
    if file_obj is None:
        return None, None, "Please upload a file"
    
    try:
        df = pd.read_excel(file_obj.name, header=23)
        
        df['Quality Check'] = ''
        df['Note'] = ''
        
        for i in range(len(df)):
            try:
                con = round(float(df.iloc[i, 9]))
                ratio_280_260 = float(df.iloc[i, 11])
                ratio_260_230 = float(df.iloc[i, 12])
                
                issues = []
                
                if con < 20:
                    issues.append("Low concentration")
                if ratio_280_260 < 1.8 or ratio_280_260 > 2.0:
                    issues.append("260/280 abnormal")
                if ratio_260_230 < 2.0:
                    issues.append("260/230 abnormal")
                
                if len(issues) > 0:
                    df.at[i, 'Quality Check'] = 'FAIL'
                    df.at[i, 'Note'] = '; '.join(issues)
                elif con >= 50 and ratio_280_260 >= 1.9 and ratio_260_230 >= 2.2:
                    df.at[i, 'Quality Check'] = 'PASS'
                    df.at[i, 'Note'] = 'Excellent quality'
                else:
                    df.at[i, 'Quality Check'] = 'ACCEPTABLE'
                    df.at[i, 'Note'] = 'Meets minimum standard'
                    
            except:
                df.at[i, 'Quality Check'] = 'ERROR'
                df.at[i, 'Note'] = 'Cannot read values'
        
        # 儲存檔案
        base_name = os.path.basename(file_obj.name)
        output_filename = f"Single_Stunner_{base_name}"
        output_path = os.path.abspath(output_filename)
        
        # 使用 openpyxl 建立有格式的 Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Stunner Results"
        
        # 定義樣式
        blue_fill = PatternFill(start_color="598BAB", fill_type="solid")
        green_fill = PatternFill(start_color="CDFFCF", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF99", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", fill_type="solid")
        gray_fill = PatternFill(start_color="D3D3D3", fill_type="solid")
        
        # 寫入數據
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # 標題列上色
        for cell in ws[1]:
            cell.fill = blue_fill
            cell.font = Font(bold=True, color="FFFFFF")
        
        # 找到 Quality Check 欄位索引
        try:
            qc_col = df.columns.get_loc("Quality Check") + 1
        except:
            qc_col = None
        
        # 為 Quality Check 欄位上色
        if qc_col:
            for row_idx in range(2, ws.max_row + 1):
                qc_cell = ws.cell(row=row_idx, column=qc_col)
                qc_value = str(qc_cell.value).upper() if qc_cell.value else ""
                
                if qc_value == 'PASS':
                    qc_cell.fill = green_fill
                elif qc_value == 'ACCEPTABLE':
                    qc_cell.fill = yellow_fill
                elif qc_value == 'FAIL':
                    qc_cell.fill = red_fill
                elif qc_value == 'ERROR':
                    qc_cell.fill = gray_fill
        
        # 設定欄寬
        for column in ws.columns:
            ws.column_dimensions[column[0].column_letter].width = 20
        
        wb.save(output_path)
        
        styled_df = style_dataframe(df)
        return styled_df, output_path, f"File loaded successfully: {base_name}"
        
    except Exception as e:
        return None, None, f"Error loading file: {str(e)}"


def load_multi_stunner(file_objs, selected_file_index):
    """
    載入多個 Stunner 檔案並支援切換瀏覽
    功能:處理多檔案上傳,允許使用者切換查看不同檔案
    """
    if not file_objs:
        return None, None, "Please upload files", [], gr.update(visible=False)
    
    file_names = [os.path.basename(f.name) for f in file_objs]
    
    if selected_file_index is None:
        selected_file_index = 0
    
    if selected_file_index >= len(file_objs):
        selected_file_index = 0
    
    try:
        df = pd.read_excel(file_objs[selected_file_index].name, header=23)
        
        df['Quality Check'] = ''
        df['Note'] = ''
        
        for i in range(len(df)):
            try:
                con = round(float(df.iloc[i, 9]))
                ratio_280_260 = float(df.iloc[i, 11])
                ratio_260_230 = float(df.iloc[i, 12])
                
                issues = []
                
                if con < 20:
                    issues.append("Low concentration")
                if ratio_280_260 < 1.8 or ratio_280_260 > 2.0:
                    issues.append("260/280 abnormal")
                if ratio_260_230 < 2.0:
                    issues.append("260/230 abnormal")
                
                if len(issues) > 0:
                    df.at[i, 'Quality Check'] = 'FAIL'
                    df.at[i, 'Note'] = '; '.join(issues)
                elif con >= 50 and ratio_280_260 >= 1.9 and ratio_260_230 >= 2.2:
                    df.at[i, 'Quality Check'] = 'PASS'
                    df.at[i, 'Note'] = 'Excellent quality'
                else:
                    df.at[i, 'Quality Check'] = 'ACCEPTABLE'
                    df.at[i, 'Note'] = 'Meets minimum standard'
                    
            except:
                df.at[i, 'Quality Check'] = 'ERROR'
                df.at[i, 'Note'] = 'Cannot read values'
        
        # 儲存當前檔案使用 openpyxl 來上色和分隔
        base_name = os.path.basename(file_names[selected_file_index])
        output_filename = f"Multi_Stunner_{base_name}"
        output_path = os.path.abspath(output_filename)
        
        # 使用 openpyxl 建立有格式的 Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Stunner Results"
        
        # 定義樣式
        blue_fill = PatternFill(start_color="598BAB", fill_type="solid")
        green_fill = PatternFill(start_color="CDFFCF", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF99", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", fill_type="solid")
        gray_fill = PatternFill(start_color="D3D3D3", fill_type="solid")
        
        # 分離原始數據和 QC 數據
        original_columns = [col for col in df.columns if col not in ['Quality Check', 'Note']]
        qc_columns = ['Quality Check', 'Note']
        
        # 寫入標題列 - 原始數據部分
        for idx, col in enumerate(original_columns, 1):
            cell = ws.cell(row=1, column=idx, value=col)
            cell.fill = blue_fill
            cell.font = Font(bold=True, color="FFFFFF")
        
        # 寫入標題列 - QC 部分從 W 欄開始 (第 23 欄)
        qc_start_col = 23  # W 欄
        for idx, col in enumerate(qc_columns, qc_start_col):
            cell = ws.cell(row=1, column=idx, value=col)
            cell.fill = blue_fill
            cell.font = Font(bold=True, color="FFFFFF")
        
        # 寫入數據
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            # 寫入原始數據
            for col_idx, col in enumerate(original_columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=row[col])
            
            # 寫入 QC 數據並上色
            quality_check = row['Quality Check']
            
            # Quality Check 欄位
            qc_cell = ws.cell(row=row_idx, column=qc_start_col, value=quality_check)
            
            # 根據 Quality Check 值上色
            if quality_check == 'PASS':
                qc_cell.fill = green_fill
            elif quality_check == 'ACCEPTABLE':
                qc_cell.fill = yellow_fill
            elif quality_check == 'FAIL':
                qc_cell.fill = red_fill
            elif quality_check == 'ERROR':
                qc_cell.fill = gray_fill
            
            # Note 欄位
            ws.cell(row=row_idx, column=qc_start_col + 1, value=row['Note'])
        
        # 設定欄寬
        for column in ws.columns:
            ws.column_dimensions[column[0].column_letter].width = 20
        
        # 儲存檔案
        wb.save(output_path)
        
        styled_df = style_dataframe(df)
        file_info = f"Viewing file {selected_file_index + 1} of {len(file_objs)}: {base_name}"
        
        return styled_df, None, file_info, file_names, gr.update(visible=True, value=output_path)
        
    except Exception as e:
        error_msg = f"Error loading file: {str(e)}"
        return None, None, error_msg, file_names, gr.update(visible=False)


# --- 3. Master Analysis System with Separated Raw Data ---
def run_master_analysis(file_objs, gel_image, mode="single"):
    """
    主分析系統 - 執行完整的品質分析流程
    功能:整合濃度分析、電泳分析,生成完整報告
    """
    if not file_objs:
        return None, None, None, None, None, "Please upload analysis files"
    
    all_results = []
    all_raw_data = []
    
    # 處理每個上傳的檔案
    for f in file_objs:
        df_raw = pd.read_excel(f.name, header=0)
        
        for i in range(len(df_raw)):
            current_sample = str(df_raw.iloc[i, 1])
            
            try:
                con = round(float(df_raw.iloc[i, 2]))
                ratio_280_260 = float(df_raw.iloc[i, 3])
                ratio_260_230 = float(df_raw.iloc[i, 4])
                
                # 保存 raw data
                raw_row = [
                    current_sample,
                    con,
                    ratio_280_260,
                    ratio_260_230
                ]
                all_raw_data.append(raw_row)
                
                # 濃度分級
                if con >= 50:
                    con_level = "High"
                elif con >= 20:
                    con_level = "Medium"
                else:
                    con_level = "Low"

                # 電泳分析
                if con >= 20:
                    if gel_image is not None:
                        smear, integrity, order_val = analyze_gel_image(gel_image.name, i+1)
                        e_val = f"{smear} / {integrity}"
                        n_val = order_val
                    else:
                        e_val = "No Gel Image"
                        n_val = "4"
                else:
                    e_val = "Concentration < 20"
                    n_val = "4"

                result_row = [
                    current_sample, 
                    con, 
                    con_level, 
                    ratio_280_260, 
                    ratio_260_230,
                    e_val, 
                    n_val
                ]
                
                all_results.append(result_row)
                
            except:
                error_row = [
                    current_sample, 
                    0, 
                    "Error", 
                    0, 
                    0, 
                    "Error", 
                    "4"
                ]
                all_results.append(error_row)

    # 建立分析結果 DataFrame
    analysis_df = pd.DataFrame(
        all_results, 
        columns=[
            "Sample Name", 
            "Concentration", 
            "Concentration Level", 
            "260/280", 
            "260/230", 
            "Electrophoresis", 
            "Order"
        ]
    )
    
    # 建立原始數據 DataFrame
    raw_data_df = pd.DataFrame(
        all_raw_data,
        columns=[
            "Sample Name",
            "Raw Concentration",
            "Raw 260/280",
            "Raw 260/230"
        ]
    )

    # 儲存到 Excel (使用 openpyxl 來上色)
    if mode == "single":
        save_path = os.path.abspath("Single_Analysis_Report.xlsx")
    else:
        save_path = os.path.abspath("Multiple_Analysis_Report.xlsx")
    
    wb = Workbook()
    
    # Sheet 1: Raw Data
    ws1 = wb.active
    ws1.title = "Raw Data"
    
    # 定義樣式
    blue_fill = PatternFill(start_color="598BAB", fill_type="solid")
    
    # 寫入 Raw Data
    for r in dataframe_to_rows(raw_data_df, index=False, header=True):
        ws1.append(r)
    
    # 標題列上色
    for cell in ws1[1]:
        cell.fill = blue_fill
        cell.font = Font(bold=True, color="FFFFFF")
    
    # 設定欄寬
    for column in ws1.columns:
        ws1.column_dimensions[column[0].column_letter].width = 20
    
    # Sheet 2: Analysis Results
    ws2 = wb.create_sheet(title="Analysis Results")
    
    # 寫入 Analysis Results
    for r in dataframe_to_rows(analysis_df, index=False, header=True):
        ws2.append(r)
    
    # 標題列上色
    for cell in ws2[1]:
        cell.fill = blue_fill
        cell.font = Font(bold=True, color="FFFFFF")
    
    # 設定欄寬
    for column in ws2.columns:
        ws2.column_dimensions[column[0].column_letter].width = 20
    
    # 儲存檔案
    wb.save(save_path)

    # 建立濃度分組表
    group_df = analysis_df[
        [
            "Sample Name", 
            "Concentration", 
            "Concentration Level", 
            "260/230"
        ]
    ].copy()
    
    group_df = group_df.sort_values(
        by=["Concentration Level", "260/230"], 
        ascending=[False, False]
    )

    # 建立定序優先順序表
    order_df = analysis_df[
        [
            "Sample Name", 
            "Order", 
            "Electrophoresis"
        ]
    ].copy()
    
    order_df = order_df.sort_values(by="Order")
    order_df['Rank'] = range(1, len(order_df) + 1)

    # 建立預覽表
    preview_df = analysis_df[
        [
            "Sample Name", 
            "Concentration", 
            "Concentration Level",
            "Order"
        ]
    ].head(10)

    return analysis_df, save_path, group_df, order_df, preview_df, "Analysis completed"


# --- 4. Password Verification ---
def check_password(password):
    """
    備註:密碼驗證函式
    """
    if password == "660531" or password == "19770531" or password == "1977531" or password == "066531" or password == "0660531" or password == "019770531" or password == "01977531" or password == "980530":
        return gr.update(visible=False), gr.update(visible=True), ""
    else:
        return gr.update(visible=True), gr.update(visible=False), "Incorrect password. Please try again."


# --- 5. Enhanced Custom CSS with Animation ---
custom_css = """
/* Global Animations */
@keyframes fadeIn {
    from { opacity: 0; transform: translateY(20px); }
    to { opacity: 1; transform: translateY(0); }
}

@keyframes slideIn {
    from { transform: translateX(-100%); opacity: 0; }
    to { transform: translateX(0); opacity: 1; }
}

@keyframes pulse {
    0%, 100% { transform: scale(1); }
    50% { transform: scale(1.05); }
}

@keyframes shimmer {
    0% { background-position: -1000px 0; }
    100% { background-position: 1000px 0; }
}

/* Container */
.gradio-container { 
    background: linear-gradient(135deg, #abbcda 20%, #cfe0f7 50%);
    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    min-height: 100vh;
    animation: fadeIn 0.8s ease-out;
}

/* Login Panel */
#login_panel { 
    background: rgba(255, 255, 255, 0.95);
    padding: 40px; 
    border-radius: 20px; 
    border: none;
    margin: auto; 
    width: 380px; 
    box-shadow: 0 20px 60px rgba(0, 0, 0, 0.3);
    backdrop-filter: blur(10px);
    animation: fadeIn 0.3s ease-out;
}

#login_panel h2 {
    color: #667eea;
    margin-bottom: 25px;
    font-weight: 700;
    letter-spacing: 1px;
}

/* Error Message */
.error-message {
    color: #e74c3c;
    font-weight: 600;
    margin-top: 10px;
    animation: fadeIn 0.3s ease-out;
}

/* Main Title */
h1 {
    background: linear-gradient(135deg, #e0e7ff 0%, #fff8e0 0%);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
    text-shadow: none !important;
    margin-bottom: 10px !important;
    font-weight: 800;
    animation: slideIn 0.8s ease-out;
}

/* Subtitle */
.subtitle {
    color: #fff8e0;
    font-size: 18px;
    margin-bottom: 30px;
    animation: fadeIn 1s ease-out;
}

/* Card Style */
.card {
    background: #e8ecfa;
    border-radius: 15px;
    padding: 25px;
    margin: 15px 0;
    box-shadow: 0 8px 20px rgba(0, 0, 0, 0.1);
    border: 1px solid rgba(102, 126, 234, 0.1);
    transition: all 0.3s ease;
    animation: fadeIn 0.6s ease-out;
}

.card:hover {
    transform: translateY(-5px);
    box-shadow: 0 12px 30px rgba(102, 126, 234, 0.2);
}

/* Primary Button */
.primary-btn { 
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%) !important;
    color: white !important; 
    font-weight: 700;
    border-radius: 12px !important;
    padding: 14px 28px !important;
    border: none !important;
    box-shadow: 0 6px 20px rgba(102, 126, 234, 0.4) !important;
    transition: all 0.3s ease !important;
    text-transform: uppercase;
    letter-spacing: 1px;
}

.primary-btn:hover {
    transform: translateY(-3px) !important;
    box-shadow: 0 10px 30px rgba(102, 126, 234, 0.5) !important;
}

.primary-btn:active {
    transform: translateY(-1px) !important;
}

/* Download Button */
.download-btn {
    background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%) !important;
    color: white !important;
    font-weight: 700;
    border-radius: 12px !important;
    padding: 14px 28px !important;
    border: none !important;
    box-shadow: 0 6px 20px rgba(17, 153, 142, 0.4) !important;
    transition: all 0.3s ease !important;
    text-transform: uppercase;
    letter-spacing: 1px;
}

.download-btn:hover {
    transform: translateY(-3px) !important;
    box-shadow: 0 10px 30px rgba(17, 153, 142, 0.5) !important;
}

/* Tab Navigation */
.tab-nav button { 
    font-weight: 600; 
    font-size: 15px;
    padding: 14px 24px !important;
    border-radius: 12px 12px 0 0 !important;
    transition: all 0.3s ease !important;
    border: none !important;
    background: rgba(255, 255, 255, 0.7) !important;
    margin-right: 5px !important;
}

.tab-nav button:hover {
    background: rgba(102, 126, 234, 0.2) !important;
    transform: translateY(-2px);
}

.tab-nav button.selected {
    background: linear-gradient(135deg, #667eea 100%, #764ba2 0%) !important;
    color: white !important;
    box-shadow: 0 4px 15px rgba(102, 126, 234, 0.3) !important;
}

/* Section Headers */
h3 { 
    color: #667eea;
    border-left: 5px solid #667eea;
    padding-left: 15px;
    margin: 25px 0 20px 0;
    font-weight: 700;
    animation: slideIn 0.5s ease-out;
}

/* Input Boxes */
.gr-box {
    border-radius: 12px !important;
    border: 2px solid #e0e7ff !important;
    transition: all 0.3s ease !important;
    background: white !important;
}

.gr-box:focus-within {
    border-color: #667eea !important;
    box-shadow: 0 0 0 4px rgba(102, 126, 234, 0.1) !important;
    transform: scale(1.02);
}

/* File Upload Area */
.file-upload {
    border: 3px dashed #667eea !important;
    border-radius: 15px !important;
    background: linear-gradient(135deg, #f8f9ff 0%, #fff 100%) !important;
    transition: all 0.3s ease !important;
    padding: 20px !important;
}

.file-upload:hover {
    background: linear-gradient(135deg, #e8ecff 0%, #f8f9ff 100%) !important;
    border-color: #764ba2 !important;
    transform: scale(1.02);
}

/* Dataframe */
.dataframe {
    border-radius: 15px !important;
    overflow: hidden !important;
    box-shadow: 0 8px 25px rgba(0, 0, 0, 0.15) !important;
    border: 1px solid rgba(102, 126, 234, 0.2) !important;
    animation: fadeIn 0.5s ease-out;
}

/* Status Box */
.gr-textbox {
    border-radius: 10px !important;
    border: 2px solid #e0e7ff !important;
    background: #f8f9ff !important;
}

/* Divider */
hr {
    border: none !important;
    border-top: 2px solid rgba(102, 126, 234, 0.2) !important;
    margin: 35px 0 !important;
}

/* Info Card */
.info-card {
    background: linear-gradient(135deg, #f8f9ff 0%, #fff 100%);
    padding: 20px;
    border-radius: 12px;
    border-left: 5px solid #667eea;
    margin: 20px 0;
    box-shadow: 0 4px 15px rgba(102, 126, 234, 0.1);
    animation: fadeIn 0.7s ease-out;
}

/* Color Legend */
.color-legend {
    background: white;
    padding: 20px;
    border-radius: 15px;
    box-shadow: 0 6px 20px rgba(0, 0, 0, 0.1);
    margin: 20px 0;
    border: 2px solid rgba(102, 126, 234, 0.2);
    animation: fadeIn 0.8s ease-out;
}

/* Radio Buttons */
.gr-radio {
    background: white !important;
    padding: 20px !important;
    border-radius: 12px !important;
    box-shadow: 0 4px 15px rgba(0, 0, 0, 0.1) !important;
    border: 2px solid rgba(102, 126, 234, 0.1) !important;
}

/* Loading Animation */
.loading {
    background: linear-gradient(90deg, #f0f0f0 25%, #e0e0e0 50%, #f0f0f0 75%);
    background-size: 1000px 100%;
    animation: shimmer 2s infinite;
}
"""


# --- 6. Gradio UI Interface ---
with gr.Blocks(title="Analysis System", css=custom_css) as demo:
    
    # === Login Interface ===
    with gr.Row(visible=True) as login_ui:
        with gr.Column(elem_id="login_panel"):
            gr.Markdown("<center><h2>ANALYSIS SYSTEM</h2></center>")
            gr.Markdown("<center><p style='color:#888;'>Enter your credentials to continue</p></center>")
            pwd = gr.Textbox(
                label="Access Key", 
                type="password", 
                placeholder="Enter your access key"
            )
            error_msg = gr.Markdown("", elem_classes="error-message")
            login_btn = gr.Button(
                "SIGN IN", 
                variant="primary", 
                elem_classes="primary-btn"
            )

    # === Main Interface ===
    with gr.Column(visible=False) as main_ui:
        gr.Markdown("<h1 style='text-align:center;'>Data Analysis Platform</h1>")
        gr.Markdown("<p class='subtitle' style='text-align:center;'>Data Quality Assessment System</p>")
        
        with gr.Tabs():
            
            # ===== Tab 0: Custom Standards Settings =====
            with gr.TabItem("Custom Standards"):
                with gr.Column(elem_classes="card"):
                    gr.Markdown("### Standards Configuration")
                    gr.Markdown("Set your own thresholds and units for quality assessment. Changes apply immediately.")
                    
                    # 檢查開關
                    with gr.Column(elem_classes="info-card"):
                        gr.Markdown("#### Enable/Disable Checks")
                        with gr.Row():
                            check_p1 = gr.Checkbox(
                                label="Check Parameter 1", 
                                value=DEFAULT_STANDARDS["check_parameter_1"]
                            )
                            check_p2 = gr.Checkbox(
                                label="Check Parameter 2", 
                                value=DEFAULT_STANDARDS["check_parameter_2"]
                            )
                            check_p3 = gr.Checkbox(
                                label="Check Parameter 3", 
                                value=DEFAULT_STANDARDS["check_parameter_3"]
                            )
                            check_img = gr.Checkbox(
                                label="Check Image Analysis", 
                                value=DEFAULT_STANDARDS["check_image"]
                            )
                    
                    gr.Markdown("---")
                    
                    with gr.Row():
                        # 參數 1
                        with gr.Column():
                            gr.Markdown("#### Parameter 1 Settings")
                            
                            p1_unit = gr.Dropdown(
                                label="Unit",
                                choices=["ng/μL", "μg/mL", "mg/mL", "ppm", "custom"],
                                value=DEFAULT_STANDARDS["parameter_1_unit"],
                                allow_custom_value=True
                            )
                            
                            p1_min = gr.Number(
                                label="Minimum Threshold",
                                value=DEFAULT_STANDARDS["parameter_1_min"],
                                info="Below this = Low quality"
                            )
                            
                            p1_pass = gr.Number(
                                label="Pass Threshold",
                                value=DEFAULT_STANDARDS["parameter_1_pass"],
                                info="Above this = High quality (PASS)"
                            )
                        
                        # 參數 2
                        with gr.Column():
                            gr.Markdown("#### Parameter 2 Settings")
                            
                            p2_unit = gr.Dropdown(
                                label="Unit",
                                choices=["ratio", "%", "score", "index", "custom"],
                                value=DEFAULT_STANDARDS["parameter_2_unit"],
                                allow_custom_value=True
                            )
                            
                            p2_min = gr.Number(
                                label="Minimum Threshold",
                                value=DEFAULT_STANDARDS["parameter_2_min"],
                                info="Normal range minimum"
                            )
                            
                            p2_max = gr.Number(
                                label="Maximum Threshold",
                                value=DEFAULT_STANDARDS["parameter_2_max"],
                                info="Normal range maximum"
                            )
                            
                            p2_pass = gr.Number(
                                label="Pass Threshold",
                                value=DEFAULT_STANDARDS["parameter_2_pass"],
                                info="For excellent quality"
                            )
                    
                    with gr.Row():
                        # 參數 3
                        with gr.Column():
                            gr.Markdown("#### Parameter 3 Settings")
                            
                            p3_unit = gr.Dropdown(
                                label="Unit",
                                choices=["ratio", "%", "score", "index", "custom"],
                                value=DEFAULT_STANDARDS["parameter_3_unit"],
                                allow_custom_value=True
                            )
                            
                            p3_min = gr.Number(
                                label="Minimum Threshold",
                                value=DEFAULT_STANDARDS["parameter_3_min"],
                                info="Normal range minimum"
                            )
                            
                            p3_pass = gr.Number(
                                label="Pass Threshold",
                                value=DEFAULT_STANDARDS["parameter_3_pass"],
                                info="For excellent quality"
                            )
                        
                        # 影像分析
                        with gr.Column():
                            gr.Markdown("#### Image Analysis Settings")
                            
                            img_unit = gr.Dropdown(
                                label="Unit",
                                choices=["intensity", "brightness", "value", "custom"],
                                value=DEFAULT_STANDARDS["image_unit"],
                                allow_custom_value=True
                            )
                            
                            img_t1 = gr.Number(
                                label="Threshold 1",
                                value=DEFAULT_STANDARDS["image_threshold_1"],
                                info="Smearing detection threshold"
                            )
                            
                            img_t2 = gr.Number(
                                label="Threshold 2",
                                value=DEFAULT_STANDARDS["image_threshold_2"],
                                info="Band visibility threshold"
                            )
                    
                    gr.Markdown("---")
                    
                    # 操作按鈕
                    with gr.Row():
                        update_btn = gr.Button(
                            "Apply Custom Standards", 
                            variant="primary", 
                            elem_classes="primary-btn",
                            scale=2
                        )
                        reset_btn = gr.Button(
                            "Reset to Default", 
                            scale=1
                        )
                    
                    standards_status = gr.Textbox(
                        label="Status", 
                        interactive=False,
                        lines=2
                    )
                    
                    gr.Markdown("---")
                    
                    # 預設檔管理
                    with gr.Column(elem_classes="info-card"):
                        gr.Markdown("#### Save/Load Presets")
                        
                        with gr.Row():
                            preset_name_input = gr.Textbox(
                                label="Preset Name",
                                placeholder="e.g., RNA_Standards, DNA_Standards"
                            )
                            save_preset_btn = gr.Button("Save Preset")
                        
                        preset_download = gr.File(
                            label="Download Preset File",
                            visible=False
                        )
                        
                        with gr.Row():
                            preset_file_input = gr.File(
                                label="Load Preset File (.json)",
                                file_count="single"
                            )
                            load_preset_btn = gr.Button("Load Preset")
                    
                    gr.Markdown("---")
                    
                    # 多檔分析區
                    with gr.Column(elem_classes="info-card"):
                        gr.Markdown("#### Batch Analysis with Custom Standards")
                        gr.Markdown("Upload multiple Excel files to analyze with current custom standards")
                        
                        custom_analysis_files = gr.File(
                            label="Upload Excel Files for Analysis",
                            file_count="multiple",
                            file_types=[".xlsx", ".xls"]
                        )
                        
                        custom_analyze_btn = gr.Button(
                            "Analyze Files",
                            variant="primary",
                            elem_classes="primary-btn"
                        )
                        
                        custom_analysis_status = gr.Textbox(
                            label="Analysis Status",
                            interactive=False,
                            lines=2
                        )
                        
                        gr.Markdown("#### Analysis Results Preview")
                        custom_analysis_preview = gr.Dataframe(
                            label="Results Preview (First 20 rows)",
                            wrap=True
                        )
                        
                        custom_analysis_download = gr.File(
                            label="Download Complete Analysis Report (Excel)",
                            visible=False
                        )
                    
                    gr.Markdown("---")
                    
                    # 當前標準顯示
                    with gr.Column(elem_classes="info-card"):
                        gr.Markdown("#### Current Standards Summary")
                        current_standards_display = gr.Textbox(
                            label="Active Standards",
                            value=json.dumps(CURRENT_STANDARDS, indent=2),
                            lines=15,
                            interactive=False
                        )
            
            # ===== Tab 1: Stunner Data Viewer =====
            with gr.TabItem("Stunner Data Viewer"):
                
                with gr.Tabs():
                    
                    # Single File Review
                    with gr.TabItem("Single File Review"):
                        with gr.Column(elem_classes="card"):
                            gr.Markdown("### Load and Quality Check Single File")
                            
                            with gr.Row():
                                with gr.Column(scale=2):
                                    stunner_file = gr.File(
                                        label="Select Stunner Excel File", 
                                        file_count="single"
                                    )
                                    with gr.Row():
                                        load_single_btn = gr.Button(
                                            "Load and Check Quality", 
                                            variant="primary", 
                                            elem_classes="primary-btn",
                                            scale=2
                                        )
                                        download_single_btn = gr.DownloadButton(
                                            "Download Results",
                                            elem_classes="download-btn",
                                            visible=False,
                                            scale=1
                                        )
                                with gr.Column(scale=1):
                                    stunner_status = gr.Textbox(
                                        label="Status Message", 
                                        interactive=False,
                                        lines=3
                                    )
                            
                            stunner_output = gr.Dataframe(
                                label="Stunner Data with Quality Annotations",
                                wrap=True
                            )
                            
                            with gr.Column(elem_classes="color-legend"):
                                gr.Markdown("""
                                **Color Legend**
                                - Green PASS: Excellent quality - Ready for sequencing
                                - Blue ACCEPTABLE: Meets minimum standard - Usable with caution
                                - Red FAIL: Does not meet requirements - Re-extraction recommended
                                - Gray ERROR: Cannot read data - Check file format
                                """)
                    
                    # Multiple Files Browser
                    with gr.TabItem("Multiple Files Browser"):
                        with gr.Column(elem_classes="card"):
                            gr.Markdown("### Browse and Compare Multiple Stunner Files")
                            
                            stunner_multi_files = gr.File(
                                label="Upload Multiple Stunner Files", 
                                file_count="multiple"
                            )
                            
                            load_multi_browser_btn = gr.Button(
                                "Load Files for Browsing", 
                                variant="primary", 
                                elem_classes="primary-btn"
                            )
                            
                            file_selector = gr.Radio(
                                label="Select File to View",
                                choices=[],
                                interactive=True
                            )
                            
                            with gr.Row():
                                multi_browser_status = gr.Textbox(
                                    label="File Information", 
                                    interactive=False,
                                    lines=2,
                                    scale=2
                                )
                                download_multi_btn = gr.DownloadButton(
                                    "Download Current File",
                                    elem_classes="download-btn",
                                    visible=False,
                                    scale=1
                                )
                            
                            stunner_multi_output = gr.Dataframe(
                                label="Selected File Data with Quality Check",
                                wrap=True
                            )
                            
                            with gr.Column(elem_classes="color-legend"):
                                gr.Markdown("""
                                **Quality Status Colors**
                                
                                PASS - ACCEPTABLE - FAIL - ERROR
                                """)
            
            # ===== Tab 2: Analysis and Grouping =====
            with gr.TabItem("Analysis and Grouping"):
                
                with gr.Tabs():
                    
                    # Single File Analysis
                    with gr.TabItem("Single File Analysis"):
                        with gr.Column(elem_classes="card"):
                            gr.Markdown("### Single File Concentration Grouping and Ratio Analysis")
                            
                            with gr.Row():
                                with gr.Column():
                                    single_analysis_file = gr.File(
                                        label="Upload Analysis File", 
                                        file_count="single"
                                    )
                                with gr.Column():
                                    single_gel_image = gr.Image(
                                        label="Upload Gel Image (Optional)", 
                                        type="filepath"
                                    )
                            
                            single_analyze_btn = gr.Button(
                                "Run Single File Analysis", 
                                variant="primary", 
                                elem_classes="primary-btn", 
                                size="lg"
                            )
                            
                            single_analysis_status = gr.Textbox(
                                label="Analysis Status", 
                                interactive=False,
                                lines=2
                            )
                            
                            gr.Markdown("---")
                            gr.Markdown("### Concentration Grouping Results")
                            
                            single_grouping_output = gr.Dataframe(
                                label="Sorted by Concentration Level and 260/230 Ratio"
                            )
                            
                            with gr.Column(elem_classes="info-card"):
                                gr.Markdown("""
                                **Grouping Criteria**
                                - High: Concentration >= 50 ng/uL
                                - Medium: 20 <= Concentration < 50 ng/uL
                                - Low: Concentration < 20 ng/uL
                                """)
                    
                    # Multiple Files Analysis
                    with gr.TabItem("Multiple Files Analysis"):
                        with gr.Column(elem_classes="card"):
                            gr.Markdown("### Multiple Files Concentration Grouping and Ratio Analysis")
                            
                            with gr.Row():
                                with gr.Column():
                                    multi_analysis_files = gr.File(
                                        label="Upload Multiple Analysis Files", 
                                        file_count="multiple"
                                    )
                                with gr.Column():
                                    multi_gel_image = gr.Image(
                                        label="Upload Gel Image (Optional)", 
                                        type="filepath"
                                    )
                            
                            multi_analyze_btn = gr.Button(
                                "Run Multiple Files Analysis", 
                                variant="primary", 
                                elem_classes="primary-btn", 
                                size="lg"
                            )
                            
                            multi_analysis_status = gr.Textbox(
                                label="Analysis Status", 
                                interactive=False,
                                lines=2
                            )
                            
                            gr.Markdown("---")
                            gr.Markdown("### Concentration Grouping Results")
                            
                            multi_grouping_output = gr.Dataframe(
                                label="Sorted by Concentration Level and 260/230 Ratio"
                            )
                            
                            with gr.Column(elem_classes="info-card"):
                                gr.Markdown("""
                                **Grouping Criteria**
                                - High: Concentration >= 50 ng/uL
                                - Medium: 20 <= Concentration < 50 ng/uL
                                - Low: Concentration < 20 ng/uL
                                """)
            
            # ===== Tab 3: Results and Download =====
            with gr.TabItem("Results and Download"):
                with gr.Column(elem_classes="card"):
                    gr.Markdown("### Complete Analysis Results and Export")
                    
                    full_analysis_output = gr.Dataframe(
                        label="Full Analysis Data"
                    )
                    
                    download_file = gr.File(
                        label="Download Complete Report (Excel)"
                    )
                    
                    with gr.Column(elem_classes="info-card"):
                        gr.Markdown("""
                        **Excel Report Structure**
                        - Section 1: Raw Data (Original measurements)
                        - Blank Row: Separator
                        - Section 2: Analysis Results (Quality assessment and sequencing order)
                        """)
            
            # ===== Tab 4: Sequencing Order =====
            with gr.TabItem("Sequencing Order"):
                with gr.Column(elem_classes="card"):
                    gr.Markdown("### Sequencing Priority Order")
                    
                    order_output = gr.Dataframe(
                        label="Sorted by Sequencing Priority"
                    )
                    
                    with gr.Column(elem_classes="info-card"):
                        gr.Markdown("""
                        **Sequencing Priority**
                        - Order 1: Highest priority (Best quality)
                        - Order 2: High priority
                        - Order 3: Medium priority
                        - Order 4: Lowest priority (Quality issues or low concentration)
                        """)
            
            # ===== Tab 5: Preview =====
            with gr.TabItem("Preview"):
                with gr.Column(elem_classes="card"):
                    gr.Markdown("### Quick Data Overview")
                    
                    preview_output = gr.Dataframe(
                        label="Key Sample Preview (Top 10)"
                    )
                    
                    with gr.Row():
                        with gr.Column(elem_classes="info-card"):
                            gr.Markdown("""
                            **Column Descriptions**
                            
                            Sample Name: Unique identifier for each sample
                            
                            Concentration: DNA/RNA concentration in ng/uL
                            
                            Concentration Level: High (>=50), Medium (20-50), Low (<20)
                            
                            Order: Sequencing priority ranking (1 = Highest, 4 = Lowest)
                            """)
                        with gr.Column(elem_classes="info-card"):
                            gr.Markdown("""
                            **Quality Thresholds**
                            
                            260/280 Ratio: Acceptable range 1.8 - 2.0
                            
                            260/230 Ratio: Acceptable >= 2.0
                            
                            Minimum Concentration: 20 ng/uL for gel analysis
                            
                            Optimal Concentration: 50 ng/uL for best results
                            """)

    # === Hidden State ===
    file_index_state = gr.State(0)
    
    # === Event Handlers ===
    
    # Login
    def handle_login(password):
        if password == "660531" or password == "19770531" or password == "1977531" or password == "066531" or password == "0660531" or password == "019770531" or password == "01977531" or password == "980530" or password == "66531":
            return gr.update(visible=False), gr.update(visible=True), ""
        else:
            return gr.update(visible=True), gr.update(visible=False), "Incorrect password. Please try again."
    
    login_btn.click(
        handle_login, 
        inputs=pwd, 
        outputs=[login_ui, main_ui, error_msg]
    )
    
    pwd.submit(
        handle_login, 
        inputs=pwd, 
        outputs=[login_ui, main_ui, error_msg]
    )
    
    # Custom Standards - Update
    update_btn.click(
        update_standards,
        inputs=[
            check_p1, check_p2, check_p3, check_img,
            p1_min, p1_pass, p1_unit,
            p2_min, p2_max, p2_pass, p2_unit,
            p3_min, p3_pass, p3_unit,
            img_t1, img_t2, img_unit
        ],
        outputs=standards_status
    ).then(
        lambda: json.dumps(CURRENT_STANDARDS, indent=2),
        outputs=current_standards_display
    )
    
    # Custom Standards - Reset
    reset_btn.click(
        reset_standards,
        outputs=[
            check_p1, check_p2, check_p3, check_img,
            p1_min, p1_pass, p1_unit,
            p2_min, p2_max, p2_pass, p2_unit,
            p3_min, p3_pass, p3_unit,
            img_t1, img_t2, img_unit,
            standards_status
        ]
    ).then(
        lambda: json.dumps(CURRENT_STANDARDS, indent=2),
        outputs=current_standards_display
    )
    
    # Custom Standards - Save Preset
    def handle_save_preset(preset_name):
        file_path, msg = save_standards_preset(preset_name)
        if file_path:
            return gr.update(value=file_path, visible=True), msg
        return gr.update(visible=False), msg
    
    save_preset_btn.click(
        handle_save_preset,
        inputs=preset_name_input,
        outputs=[preset_download, standards_status]
    )
    
    # Custom Standards - Load Preset
    load_preset_btn.click(
        load_standards_preset,
        inputs=preset_file_input,
        outputs=[
            check_p1, check_p2, check_p3, check_img,
            p1_min, p1_pass, p1_unit,
            p2_min, p2_max, p2_pass, p2_unit,
            p3_min, p3_pass, p3_unit,
            img_t1, img_t2, img_unit,
            standards_status
        ]
    ).then(
        lambda: json.dumps(CURRENT_STANDARDS, indent=2),
        outputs=current_standards_display
    )
    
    # Custom Standards - Batch Analysis
    def handle_custom_analysis(files):
        if not files:
            return None, gr.update(visible=False), "Please upload Excel files"
        
        df, download_path, msg = analyze_custom_standards_data(files)
        
        if df is not None:
            return df.head(20), gr.update(value=download_path, visible=True), msg
        else:
            return None, gr.update(visible=False), msg
    
    custom_analyze_btn.click(
        handle_custom_analysis,
        inputs=custom_analysis_files,
        outputs=[custom_analysis_preview, custom_analysis_download, custom_analysis_status]
    )
    
    # Single File Load
    def handle_single_load(file_obj):
        df, output_path, msg = load_single_stunner(file_obj)
        if df is not None:
            return df, msg, gr.update(visible=True, value=output_path)
        return df, msg, gr.update(visible=False)
    
    load_single_btn.click(
        handle_single_load,
        inputs=stunner_file,
        outputs=[stunner_output, stunner_status, download_single_btn]
    )
    
    # Multiple Files Browser
    def handle_multi_load(files):
        if not files:
            return None, None, "Please upload files", gr.update(choices=[]), gr.update(visible=False)
        
        file_names = [os.path.basename(f.name) for f in files]
        df, _, msg, _, download_btn = load_multi_stunner(files, 0)
        
        return df, None, msg, gr.update(choices=file_names, value=file_names[0]), download_btn
    
    load_multi_browser_btn.click(
        handle_multi_load,
        inputs=stunner_multi_files,
        outputs=[stunner_multi_output, file_index_state, multi_browser_status, file_selector, download_multi_btn]
    )
    
    def handle_file_selection(files, selected_name):
        if not files or not selected_name:
            return None, "No file selected", gr.update(visible=False)
        
        file_names = [os.path.basename(f.name) for f in files]
        if selected_name in file_names:
            idx = file_names.index(selected_name)
            df, _, msg, _, download_btn = load_multi_stunner(files, idx)
            return df, msg, download_btn
        return None, "File not found", gr.update(visible=False)
    
    file_selector.change(
        handle_file_selection,
        inputs=[stunner_multi_files, file_selector],
        outputs=[stunner_multi_output, multi_browser_status, download_multi_btn]
    )
    
    # Single File Analysis
    def handle_single_analysis(file_obj, gel_img):
        if file_obj is None:
            return None, None, None, None, None, "Please upload a file"
        result = run_master_analysis([file_obj], gel_img, mode="single")
        return result
    
    single_analyze_btn.click(
        handle_single_analysis,
        inputs=[single_analysis_file, single_gel_image],
        outputs=[
            full_analysis_output,
            download_file,
            single_grouping_output,
            order_output,
            preview_output,
            single_analysis_status
        ]
    )
    
    # Multiple Files Analysis
    def handle_multi_analysis(files, gel_img):
        if not files:
            return None, None, None, None, None, "Please upload files"
        result = run_master_analysis(files, gel_img, mode="multiple")
        return result
    
    multi_analyze_btn.click(
        handle_multi_analysis,
        inputs=[multi_analysis_files, multi_gel_image],
        outputs=[
            full_analysis_output,
            download_file,
            multi_grouping_output,
            order_output,
            preview_output,
            multi_analysis_status
        ]
    )


if __name__ == "__main__":
    demo.launch(
        share=False, 
        server_name="127.0.0.1", 
        server_port=7860
    )
